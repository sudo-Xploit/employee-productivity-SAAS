import io
import os
import uuid
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any, BinaryIO

import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.models.department import Department
from app.models.employee import Employee
from app.models.project import Project
from app.models.timesheet import Timesheet
from app.services import analytics_service


def create_report_directory() -> str:
    """Create directory for storing reports if it doesn't exist."""
    reports_dir = os.path.join(os.getcwd(), "reports")
    if not os.path.exists(reports_dir):
        os.makedirs(reports_dir)
    return reports_dir


def generate_pdf_report(db: Session) -> Tuple[str, str]:
    """Generate a PDF report with department ROI summary, employee productivity, and insights."""
    # Create reports directory
    reports_dir = create_report_directory()
    
    # Generate unique filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"productivity_report_{timestamp}.pdf"
    filepath = os.path.join(reports_dir, filename)
    
    # Create PDF document
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = styles["Title"]
    heading_style = styles["Heading2"]
    normal_style = styles["Normal"]
    
    # Add title
    elements.append(Paragraph("Employee Productivity and Cost Analysis Report", title_style))
    elements.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Spacer(1, 0.25 * inch))
    
    # Company Overview
    company_data = analytics_service.get_company_analytics(db)
    elements.append(Paragraph("Company Overview", heading_style))
    
    company_table_data = [
        ["Metric", "Value"],
        ["Total Employees", str(company_data["employee_count"])],
        ["Total Departments", str(company_data["department_count"])],
        ["Total Projects", str(company_data["project_count"])],
        ["Total Revenue", f"${company_data['total_revenue']:,.2f}"],
        ["Total Cost", f"${company_data['total_cost']:,.2f}"],
        ["Profit", f"${company_data['profit']:,.2f}"],
        ["Profit Margin", f"{company_data['profit_margin']:.2f}%"],
        ["ROI", f"{company_data['roi']:.2f}"],
        ["Productivity Index", f"${company_data['productivity_index']:.2f}/hour"],
    ]
    
    company_table = Table(company_table_data, colWidths=[2.5*inch, 2.5*inch])
    company_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (1, 0), 12),
        ('BACKGROUND', (0, 1), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
    ]))
    elements.append(company_table)
    elements.append(Spacer(1, 0.25 * inch))
    
    # Department ROI Summary
    elements.append(Paragraph("Department ROI Summary", heading_style))
    
    departments = db.query(Department).all()
    dept_data = [["Department", "Budget", "Revenue", "Cost", "ROI", "Productivity"]]
    
    for dept in departments:
        analytics = analytics_service.get_department_analytics(db, dept.id)
        dept_data.append([
            dept.name,
            f"${dept.budget:,.2f}",
            f"${analytics['total_revenue']:,.2f}",
            f"${analytics['total_salary_cost'] + analytics['total_project_cost']:,.2f}",
            f"{analytics['roi']:.2f}",
            f"${analytics['productivity_index']:.2f}/hour"
        ])
    
    dept_table = Table(dept_data, colWidths=[1.2*inch, 1*inch, 1*inch, 1*inch, 0.8*inch, 1*inch])
    dept_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
    ]))
    elements.append(dept_table)
    elements.append(Spacer(1, 0.25 * inch))
    
    # Top Performing Employees
    elements.append(Paragraph("Top Performing Employees", heading_style))
    
    top_employees = analytics_service.get_top_performers(db, limit=10)
    employee_data = [["Employee", "Department", "Revenue", "Salary", "ROI"]]
    
    for emp in top_employees:
        employee_data.append([
            emp["employee_name"],
            emp["department_name"],
            f"${emp['revenue_generated']:,.2f}",
            f"${emp['salary']:,.2f}",
            f"{(emp['revenue_generated'] - emp['salary']) / emp['salary']:.2f}"
        ])
    
    emp_table = Table(employee_data, colWidths=[1.5*inch, 1.5*inch, 1*inch, 1*inch, 1*inch])
    emp_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),
    ]))
    elements.append(emp_table)
    elements.append(Spacer(1, 0.25 * inch))
    
    # AI Insights
    elements.append(Paragraph("AI Insights", heading_style))
    
    insights = []
    
    # Company-level insights
    if company_data["roi"] < 0:
        insights.append("The company is currently operating at a loss. Review department budgets and project allocations.")
    
    if company_data["budget_utilization"] > 90:
        insights.append("The company is approaching its overall budget limit. Consider cost-cutting measures or budget adjustments.")
    
    if company_data["productivity_index"] < 50 and company_data["total_hours"] > 0:
        insights.append("Overall productivity is below target. Consider training programs or process improvements.")
    
    # Department-level insights
    low_roi_depts = []
    high_roi_depts = []
    
    for dept in departments:
        analytics = analytics_service.get_department_analytics(db, dept.id)
        if analytics.get('roi', 0) < 0:
            low_roi_depts.append(dept.name)
        if analytics.get('roi', 0) > 1.0:  # ROI > 100%
            high_roi_depts.append(dept.name)
    
    if low_roi_depts:
        insights.append(f"The following departments have negative ROI: {', '.join(low_roi_depts)}. Review their operations and resource allocation.")
    
    if high_roi_depts:
        insights.append(f"The following departments have excellent ROI: {', '.join(high_roi_depts)}. Consider studying their practices for company-wide implementation.")
    
    # Add insights to PDF
    for insight in insights:
        elements.append(Paragraph(f"• {insight}", normal_style))
        elements.append(Spacer(1, 0.1 * inch))
    
    # Build PDF
    doc.build(elements)
    
    # Save PDF to file
    with open(filepath, "wb") as f:
        f.write(buffer.getvalue())
    
    return filename, filepath


def generate_excel_report(db: Session) -> Tuple[str, str]:
    """Generate an Excel report with multiple sheets for different analytics."""
    # Create reports directory
    reports_dir = create_report_directory()
    
    # Generate unique filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"productivity_report_{timestamp}.xlsx"
    filepath = os.path.join(reports_dir, filename)
    
    # Create workbook
    wb = Workbook()
    
    # Company Overview Sheet
    company_sheet = wb.active
    company_sheet.title = "Company Overview"
    
    company_data = analytics_service.get_company_analytics(db)
    
    # Add title
    company_sheet["A1"] = "Company Overview"
    company_sheet["A1"].font = Font(size=14, bold=True)
    company_sheet.merge_cells("A1:B1")
    
    # Add data
    metrics = [
        ["Total Employees", company_data["employee_count"]],
        ["Total Departments", company_data["department_count"]],
        ["Total Projects", company_data["project_count"]],
        ["Total Revenue", company_data["total_revenue"]],
        ["Total Cost", company_data["total_cost"]],
        ["Profit", company_data["profit"]],
        ["Profit Margin", f"{company_data['profit_margin']:.2f}%"],
        ["ROI", company_data["roi"]],
        ["Productivity Index", company_data["productivity_index"]],
        ["Total Hours", company_data["total_hours"]],
    ]
    
    for i, (metric, value) in enumerate(metrics, start=3):
        company_sheet[f"A{i}"] = metric
        company_sheet[f"B{i}"] = value
        if isinstance(value, float):
            company_sheet[f"B{i}"].number_format = '#,##0.00'
    
    # Format columns
    for col in ["A", "B"]:
        company_sheet.column_dimensions[col].width = 20
    
    # Department Sheet
    dept_sheet = wb.create_sheet(title="Departments")
    
    # Add headers
    headers = ["Department", "Budget", "Revenue", "Cost", "Profit", "ROI", "Productivity"]
    for col_num, header in enumerate(headers, start=1):
        cell = dept_sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add department data
    departments = db.query(Department).all()
    for row_num, dept in enumerate(departments, start=2):
        analytics = analytics_service.get_department_analytics(db, dept.id)
        
        dept_sheet.cell(row=row_num, column=1, value=dept.name)
        dept_sheet.cell(row=row_num, column=2, value=dept.budget).number_format = '#,##0.00'
        dept_sheet.cell(row=row_num, column=3, value=analytics["total_revenue"]).number_format = '#,##0.00'
        dept_sheet.cell(row=row_num, column=4, value=analytics["total_salary_cost"] + analytics["total_project_cost"]).number_format = '#,##0.00'
        dept_sheet.cell(row=row_num, column=5, value=analytics["profit"]).number_format = '#,##0.00'
        dept_sheet.cell(row=row_num, column=6, value=analytics["roi"]).number_format = '0.00'
        dept_sheet.cell(row=row_num, column=7, value=analytics["productivity_index"]).number_format = '0.00'
    
    # Format columns
    for col in range(1, len(headers) + 1):
        dept_sheet.column_dimensions[get_column_letter(col)].width = 15
    
    # Employees Sheet
    emp_sheet = wb.create_sheet(title="Employees")
    
    # Add headers
    headers = ["Employee", "Department", "Salary", "Revenue", "ROI", "Productivity", "Utilization"]
    for col_num, header in enumerate(headers, start=1):
        cell = emp_sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add employee data
    employees = db.query(Employee).all()
    for row_num, emp in enumerate(employees, start=2):
        analytics = analytics_service.get_employee_analytics(db, emp.id)
        department = db.query(Department).filter(Department.id == emp.department_id).first()
        
        emp_sheet.cell(row=row_num, column=1, value=emp.name)
        emp_sheet.cell(row=row_num, column=2, value=department.name if department else "")
        emp_sheet.cell(row=row_num, column=3, value=emp.salary).number_format = '#,##0.00'
        emp_sheet.cell(row=row_num, column=4, value=emp.revenue_generated).number_format = '#,##0.00'
        emp_sheet.cell(row=row_num, column=5, value=analytics["roi"]).number_format = '0.00'
        emp_sheet.cell(row=row_num, column=6, value=analytics["productivity_index"]).number_format = '0.00'
        emp_sheet.cell(row=row_num, column=7, value=analytics["utilization_rate"]).number_format = '0.00%'
    
    # Format columns
    for col in range(1, len(headers) + 1):
        emp_sheet.column_dimensions[get_column_letter(col)].width = 15
    
    # Projects Sheet
    proj_sheet = wb.create_sheet(title="Projects")
    
    # Add headers
    headers = ["Project", "Department", "Cost", "Revenue", "Profit", "Margin", "Hours"]
    for col_num, header in enumerate(headers, start=1):
        cell = proj_sheet.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    
    # Add project data
    projects = db.query(Project).all()
    for row_num, proj in enumerate(projects, start=2):
        analytics = analytics_service.get_project_analytics(db, proj.id)
        department = db.query(Department).filter(Department.id == proj.department_id).first()
        
        proj_sheet.cell(row=row_num, column=1, value=proj.name)
        proj_sheet.cell(row=row_num, column=2, value=department.name if department else "")
        proj_sheet.cell(row=row_num, column=3, value=analytics["total_cost"]).number_format = '#,##0.00'
        proj_sheet.cell(row=row_num, column=4, value=proj.revenue).number_format = '#,##0.00'
        proj_sheet.cell(row=row_num, column=5, value=analytics["profit"]).number_format = '#,##0.00'
        proj_sheet.cell(row=row_num, column=6, value=analytics["profit_margin"] / 100).number_format = '0.00%'
        proj_sheet.cell(row=row_num, column=7, value=analytics["total_hours"]).number_format = '#,##0.00'
    
    # Format columns
    for col in range(1, len(headers) + 1):
        proj_sheet.column_dimensions[get_column_letter(col)].width = 15
    
    # Save workbook
    wb.save(filepath)
    
    return filename, filepath
